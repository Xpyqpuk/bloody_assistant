from datetime import datetime, date
from time import strptime
from openpyxl import load_workbook

wb = load_workbook('data.xlsx', data_only=True)
ws = wb.active

donations_dates_list = [i.value for i in ws["A"][1:] if i.value is not None]

print(ws.max_row)

NEW_FORMAT = '%d-%m-%Y'
last_date = ws.cell(row=ws.max_row, column=1).value
print(last_date)

def list_all_dates(worksheet):
    donations_dates_string = '\n'.join([i.strftime(NEW_FORMAT) for i in donations_dates_list])
    return donations_dates_string

def get_info_by_date(date):
    position = 0
    info = []
    for i in range(len(donations_dates_list)):
        if donations_dates_list[i].strftime(NEW_FORMAT) == date:
            position = i + 2
    if not position: return info
    info = [i.value for i in ws[position]]
    info.pop(0)
    first = f"Days from last donation - {info[0]}\n"
    second = f"Involved hand - {'right' if info[1] else 'left'}\n"
    third = f"Hemoglobin(g/l) - {info[2]}\n"
    fourth = f"Hematocrit(%) - {info[3]}\n"
    fifth = f"Platelets(10^9 cells/l) - {info[4]}\n"
    sixth = f"White blood cells(10^9 cells/l) - {info[5]}\n"
    seventh = f"Result of donation(idk) - {info[6]}\n"
    result_string = f"Here's all I could find for date {date}:\n\n"+first+second+third+fourth+fifth+sixth+seventh
    
    return result_string


input_string = '23-12-2023,1,156,42.5,244,4.2,210'
def add_info(input_string):
    arr = input_string.split(',')
    arr[0] = datetime.strptime(arr[0],NEW_FORMAT)
    arr.insert(1, (arr[0] - last_date).days)
    arr[2] = int(arr[2])
    arr[3] = int(arr[3])
    arr[4] = float(arr[4])
    arr[5] = int(arr[5])
    arr[6] = float(arr[6])
    arr[7] = int(arr[7])
    ws.append(arr)
    wb.save('data.xlsx')
    first = f"Days from last donation - {arr[1]}\n"
    second = f"Involved hand - {'right' if arr[2] else 'left'}\n"
    third = f"Hemoglobin(g/l) - {arr[3]}\n"
    fourth = f"Hematocrit(%) - {arr[4]}\n"
    fifth = f"Platelets(10^9 cells/l) - {arr[5]}\n"
    sixth = f"White blood cells(10^9 cells/l) - {arr[6]}\n"
    seventh = f"Result of donation(idk) - {arr[7]}\n"
    result_string = f"Values:\n\nDate - {arr[0].strftime(NEW_FORMAT)}\n"+first+second+third+fourth+fifth+sixth+seventh+"\nsuccessfully added, anything else?"
    
    return result_string

# print(add_info(input_string))
# print([i.value for i in ws['A19':'H19'][0]])


# print(get_info_by_date('03-10-2023'))
# print(list_all_dates(ws))